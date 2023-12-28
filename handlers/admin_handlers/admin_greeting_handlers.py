import os
import sqlite3

import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ParseMode
from loguru import logger

from database.database import recording_data_of_users_who_launched_the_bot
from keyboards.admin_keyboards.admin_keyboards import admin_create_greeting_keyboard
from system.dispatcher import bot, dp


@dp.message_handler(commands=['admin_start'])
async def admin_send_start(message: types.Message, state: FSMContext):
    """Обработчик команды /start, он же пост приветствия 👋"""
    await state.finish()  # Завершаем текущее состояние машины состояний
    await state.reset_state()  # Сбрасываем все данные машины состояний, до значения по умолчанию

    # Получаем информацию о пользователе
    user_id = message.from_user.id
    username = message.from_user.username
    first_name = message.from_user.first_name
    last_name = message.from_user.last_name
    join_date = message.date.strftime("%Y-%m-%d %H:%M:%S")

    logger.info(f"Пользователь {username} ({user_id}) запустил бота в {join_date}")
    # Записываем информацию о пользователе в базу данных
    recording_data_of_users_who_launched_the_bot(user_id, username, first_name, last_name, join_date)

    greeting_keyboard = admin_create_greeting_keyboard()
    data = (f"<b>Привет админ {first_name} {last_name}, спасибо что поддерживаешь на нашего бота 🤖!</b>\n\n"
            f"Для запуска админ панели нажми на /admin_start")
    await bot.send_message(message.from_user.id, text=data, reply_markup=greeting_keyboard, parse_mode=ParseMode.HTML)


# Функция для создания файла Excel с данными заказов
def create_excel_file(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'Имя'
    # sheet['C1'] = 'Фамилия'
    # sheet['D1'] = 'Город'
    sheet['C1'] = 'Номер телефона'
    sheet['D1'] = 'Дата регистрации'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # Имя
        sheet.cell(row=index, column=3).value = order[2]  # Фамилия
        sheet.cell(row=index, column=4).value = order[3]  # Город
        # sheet.cell(row=index, column=5).value = order[4]  # Номер телефона
        # sheet.cell(row=index, column=6).value = order[5]  # Дата регистрации

    return workbook


@dp.callback_query_handler(lambda c: c.data == 'get_a_list_of_users_registered_in_the_bot')
async def export_data(message: types.Message, state: FSMContext):
    """Получение списка зарегистрированных пользователей"""
    await state.finish()  # Завершаем текущее состояние машины состояний
    await state.reset_state()  # Сбрасываем все данные машины состояний, до значения по умолчанию
    try:
        # if message.from_user.id not in [535185511, 301634256]:
        #     await message.reply('У вас нет доступа к этой команде.')
        #     return

        conn = sqlite3.connect('your_database.db')  # Подключение к базе данных SQLite
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users")  # Получение данных из базы данных
        orders = cursor.fetchall()
        workbook = create_excel_file(orders)  # Создание файла Excel
        filename = 'Зарегистрированные пользователи в боте.xlsx'
        workbook.save(filename)  # Сохранение файла
        with open(filename, 'rb') as file:
            text = ("Данные пользователей зарегистрированных в боте\n\n"
                    "Для запуска админ панели или возврата в начальное меню нажми на /start_admin")
            await bot.send_document(message.from_user.id, document=file, caption=text)  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(e)


def create_excel_file_start(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'username'
    sheet['C1'] = 'Имя'
    sheet['D1'] = 'Фамилия'
    sheet['E1'] = 'Дата запуска бота'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # username
        sheet.cell(row=index, column=3).value = order[2]  # Имя
        sheet.cell(row=index, column=4).value = order[3]  # Фамилия
        sheet.cell(row=index, column=5).value = order[4]  # Дата запуска бота

    return workbook


@dp.callback_query_handler(lambda c: c.data == 'get_users_who_launched_the_bot')
async def get_users_who_launched_the_bot(message: types.Message, state: FSMContext):
    """Получение данных пользователей, запускающих бота"""
    await state.finish()  # Завершаем текущее состояние машины состояний
    await state.reset_state()  # Сбрасываем все данные машины состояний, до значения по умолчанию
    try:
        # if message.from_user.id not in [535185511, 301634256]:
        #     await message.reply('У вас нет доступа к этой команде.')
        #     return
        # Подключение к базе данных SQLite
        conn = sqlite3.connect('your_database.db')
        cursor = conn.cursor()
        # Получение данных из базы данных
        cursor.execute("SELECT * FROM users_start")
        orders = cursor.fetchall()
        # Создание файла Excel
        workbook = create_excel_file_start(orders)
        filename = 'Данные пользователей запустивших бота.xlsx'
        workbook.save(filename)  # Сохранение файла
        with open(filename, 'rb') as file:
            text = ("Данные пользователей запустивших бота\n\n"
                    "Для запуска админ панели или возврата в начальное меню нажми на /start_admin")
            await bot.send_document(message.from_user.id, document=file, caption=text)  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(e)


class MyStates(StatesGroup):
    waiting_for_message = State()
    waiting_for_image = State()
    waiting_for_caption = State()


@dp.callback_query_handler(lambda c: c.data == 'send_an_image_to_bot_users')
async def send_an_image_to_bot_users(message: types.Message, state: FSMContext):
    """Запрашивает изображение у администратора"""
    await state.finish()  # Завершаем текущее состояние машины состояний
    await state.reset_state()  # Сбрасываем все данные машины состояний, до значения по умолчанию
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        await bot.send_message(message.from_user.id, text="Загрузите изображение для рассылки:")
        await MyStates.waiting_for_image.set()  # Устанавливаем состояние "ожидание изображения"
    except Exception as e:
        logger.error(e)


@dp.message_handler(state=MyStates.waiting_for_image, content_types=types.ContentType.PHOTO)
async def process_send_image(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать загруженного изображения и переходить в состояние "ожидание подписи"
    """

    await state.update_data(photo=message.photo[-1].file_id)
    await bot.send_message(message.from_user.id, text="Введите подпись к изображению:")
    await MyStates.waiting_for_caption.set()


@dp.message_handler(state=MyStates.waiting_for_caption, content_types=types.ContentType.TEXT)
async def process_send_image_with_caption(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать введенной подписи и выполнять рассылку
    """
    async with state.proxy() as data:
        data['caption'] = message.text
    # Получаем список уникальных ID пользователей из базы данных
    user_ids = get_user_ids()
    if user_ids:
        # Рассылка изображения с подписью всем пользователям из списка
        for user_id in user_ids:
            try:
                # Отправляем изображение с подписью
                await bot.send_photo(user_id, data['photo'], caption=data['caption'])
            except Exception as e:
                print(f"Ошибка при отправке изображения с подписью пользователю {user_id}: {str(e)}")
    await message.answer("Изображение успешно разослано всем пользователям.")
    await state.finish()


@dp.callback_query_handler(lambda c: c.data == 'send_a_message_to_bot_users')
async def send_a_message_to_bot_users(message: types.Message, state: FSMContext):
    """Запрашивает текст сообщения у администратора"""
    await state.finish()  # Завершаем текущее состояние машины состояний
    await state.reset_state()  # Сбрасываем все данные машины состояний, до значения по умолчанию
    try:
        if message.from_user.id not in [535185511, 301634256]:
            await message.reply('У вас нет доступа к этой команде.')
            return
        await bot.send_message(message.from_user.id, text="Введите текст для рассылки:")
        await MyStates.waiting_for_message.set()  # Устанавливаем состояние "ожидание сообщения"
    except Exception as e:
        logger.error(e)


@dp.message_handler(state=MyStates.waiting_for_message, content_types=types.ContentType.TEXT)
async def process_send_message(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать введенного текста и выполнять рассылку
    """
    async with state.proxy() as data:
        data['message_text'] = message.text
    # Получаем список уникальных ID пользователей из базы данных
    user_ids = get_user_ids()
    if user_ids:
        # Рассылка сообщения всем пользователям из списка
        for user_id in user_ids:
            try:
                await bot.send_message(user_id, data['message_text'], parse_mode=ParseMode.MARKDOWN)
            except Exception as e:
                print(f"Ошибка при отправке сообщения пользователю {user_id}: {str(e)}")
    await message.answer("Сообщение успешно разослано всем пользователям.")
    await state.finish()


def get_user_ids():
    """Получаем уникальные ID пользователей из базы данных"""
    try:
        conn = sqlite3.connect('your_database.db')  # Замените 'your_database.db' на имя вашей базы данных
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT user_id FROM users_start")
        user_ids = [row[0] for row in cursor.fetchall()]
        return user_ids
    except Exception as e:
        print(f"Ошибка при получении ID пользователей из базы данных: {str(e)}")
        return []


def create_excel_file_email(orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    sheet['A1'] = 'ID аккаунта пользователя'
    sheet['B1'] = 'username'
    sheet['C1'] = 'email'
    sheet['D1'] = 'Дата регистрации'
    # Заполнение данными заказов
    for index, order in enumerate(orders, start=2):
        sheet.cell(row=index, column=1).value = order[0]  # ID аккаунта пользователя
        sheet.cell(row=index, column=2).value = order[1]  # username
        sheet.cell(row=index, column=3).value = order[2]  # email
        sheet.cell(row=index, column=4).value = order[3]  # Дата регистрации

    return workbook


@dp.callback_query_handler(lambda c: c.data == 'get_a_email')
async def get_a_email(message: types.Message, state: FSMContext):
    """Получение данных пользователей, запускающих бота"""
    await state.finish()  # Завершаем текущее состояние машины состояний
    await state.reset_state()  # Сбрасываем все данные машины состояний, до значения по умолчанию
    try:
        # if message.from_user.id not in [535185511, 301634256]:
        #     await message.reply('У вас нет доступа к этой команде.')
        #     return
        # Подключение к базе данных SQLite
        conn = sqlite3.connect('your_database.db')
        cursor = conn.cursor()
        # Получение данных из базы данных
        cursor.execute("SELECT * FROM users_email")
        orders = cursor.fetchall()
        # Создание файла Excel
        workbook = create_excel_file_email(orders)
        filename = 'Данные пользователей записавших email.xlsx'
        workbook.save(filename)  # Сохранение файла
        with open(filename, 'rb') as file:
            text = ("Данные пользователей запустивших бота\n\n"
                    "Для запуска админ панели или возврата в начальное меню нажми на /start_admin")
            await bot.send_document(message.from_user.id, document=file, caption=text)  # Отправка файла пользователю
        os.remove(filename)  # Удаление файла
    except Exception as e:
        logger.error(e)


def register_admin_greeting_handler():
    """Регистрируем handlers для бота"""
    dp.register_message_handler(admin_send_start)
    dp.register_message_handler(export_data)
    dp.register_message_handler(get_users_who_launched_the_bot)
    dp.register_message_handler(send_a_message_to_bot_users)
    dp.register_message_handler(send_an_image_to_bot_users)
    dp.register_message_handler(get_a_email)
